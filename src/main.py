import flet as ft
from logic.dbManager import *
from logic.config import *
from logic.notification_helper import *
from dateutil.parser import parse
import asyncio
from openpyxl import load_workbook

def main(page: ft.Page):
    # Cargar configuración y la base de datos
    config = load_config()
    db_initialize()
    
    def theme_changed(e):
        page.theme_mode = ft.ThemeMode.LIGHT if e.control.value == "light" else ft.ThemeMode.DARK
        page.update()
        save_settings(None)

    page.theme_mode = ft.ThemeMode.LIGHT if config["theme"] == "light" else ft.ThemeMode.DARK
    page.title = "Simple Gym"
    def resized_window(e):
        #dashboard view
        dashboard_view.height = page.window.height - 100
        dashboard_view.width = page.window.width - 300
        card_active.width = (page.window.width - 400) / 4
        card_income.width = (page.window.width - 400) / 4
        card_clients.width = (page.window.width - 400) / 4
        card_pendent.width = (page.window.width - 400) / 4
        clients_chart.width = (page.window.width - 400) / 2
        clients_chart.height = (page.window.height - 200) / 2
        income_chart.width = (page.window.width - 400) / 2
        income_chart.height = (page.window.height - 200) / 2
        #client view
        search_input.width = page.window.width - 300
        clientes.width = page.window.width -200
        #clientes.height = page.window.height - 250
        responsive_row.width = page.window.width -200
        responsive_column.height = page.window.height -260
        #settings view
        settings_view.height = page.window.height - 100
        page.update()
    page.on_resized = resized_window
    if not active_application():
        is_trial = check_trial()
        if is_trial:
            page.title = page.title + " - Sem Registrar"
        else:    
            dialog_act = ft.AlertDialog(modal=True,title=ft.Text("O periodo de prova Acabou", size=24, weight="bold", color=ft.Colors.RED),
                                         content=ft.Column(width=300,height=300,controls=[
                                         ft.Row(alignment=ft.MainAxisAlignment.CENTER,controls=[ft.Icon(name=ft.Icons.ALARM,size=100,color=ft.Colors.RED)]),
                                         ft.Text("-Para continuar utilizando Simple Gym para gerenciar sua academia você precisa introducir uma clave de Ativação ",size=15)]),
                                         )
            page.overlay.append(dialog_act)
            dialog_act.open = True

    # Views do Navigation Rail
    #region DASHBOARD VIEW

    income_data = None
    total_clients = None
    active_clients = None
    pendent_clients = None

    dd = ft.Dropdown(label='Período', options=[
        ft.DropdownOption('1', 'Mensal'),
        ft.DropdownOption('2', 'Trimestral'),
        ft.DropdownOption('3', 'Anual')
    ], on_change=lambda e: load_dashboard_data(e.control.value))

    card_income = ft.Card(color='GREEN', width=200, height=130, content=ft.Column(controls=[
        ft.Row(controls=[
            ft.Icon(ft.Icons.ATTACH_MONEY, color='WHITE', size=50),
            ft.Container(expand=True, height=50),
            ft.Text(str(income_data), color='WHITE', size=50)]),
        ft.Row(alignment=ft.MainAxisAlignment.END, controls=[
            ft.Text(color='WHITE', value='GANHOS TOTAIS', size=20)
        ])
    ]))
    card_clients = ft.Card(color='PURPLE', width=200, height=130, content=ft.Column(controls=[
        ft.Row(controls=[
            ft.Icon(ft.Icons.GROUP, color='WHITE', size=50),
            ft.Container(expand=True, height=50),
            ft.Text(total_clients, color='WHITE', size=50)]),
        ft.Row(alignment=ft.MainAxisAlignment.END, controls=[
            ft.Text(color='WHITE', value='CLIENTES', size=20)
        ])
    ]))

    card_pendent = ft.Card(color='RED', width=200, height=130, content=ft.Column(controls=[
        ft.Row(controls=[
            ft.Icon(ft.Icons.PENDING_ACTIONS, color='WHITE', size=50),
            ft.Container(expand=True, height=50),
            ft.Text(str(pendent_clients), color='WHITE', size=50)]),
        ft.Row(alignment=ft.MainAxisAlignment.END, controls=[
            ft.Text(color='WHITE', value='PENDENTES', size=20)
        ])
    ]))

    card_active = ft.Card(color='BLUE', width=200, height=130, content=ft.Column(controls=[
        ft.Row(controls=[
            ft.Icon(ft.Icons.CHECK_CIRCLE, color='WHITE', size=50),
            ft.Container(expand=True, height=50),
            ft.Text(str(active_clients), color='WHITE', size=50)]),
        ft.Row(alignment=ft.MainAxisAlignment.END, controls=[
            ft.Text(color='WHITE', value='CLIENTES ATIVOS', size=20)
        ])
    ]))

    normal_title_style = ft.TextStyle(size=16, color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD)
    hover_title_style = ft.TextStyle(size=22,color=ft.Colors.WHITE,weight=ft.FontWeight.BOLD,shadow=ft.BoxShadow(blur_radius=2, color=ft.Colors.BLACK54),)
    normal_radius = 50
    hover_radius = 60
    #charts
    def on_chart_event(e):
        for idx, section in enumerate(clients_chart.sections):
            if idx == e.section_index:
                section.radius = hover_radius
                section.title_style = hover_title_style
            else:
                section.radius = normal_radius
                section.title_style = normal_title_style
        clients_chart.update()
    clients_chart = ft.PieChart(width=400,height=400,sections_space=0,center_space_radius=40,on_chart_event=on_chart_event)
    income_chart = ft.BarChart(
        width=400,
        height=400,
        bar_groups=[],  # Inicialmente vazio
        border=ft.border.all(1, ft.Colors.GREY_400),
        left_axis=ft.ChartAxis(
            labels_size=40,
            title=ft.Text("Ingresos por mes"),
            title_size=40
        ),
        bottom_axis=ft.ChartAxis(
            title=ft.Text("Meses"),
            labels=[],  # Inicialmente vazio
            labels_size=40,
        ),
        horizontal_grid_lines=ft.ChartGridLines(
            color=ft.Colors.GREY_300,
            width=1,
            dash_pattern=[3, 3]
        ),
        tooltip_bgcolor=ft.Colors.with_opacity(0.5, ft.Colors.GREY_300),
        max_y=100,  # Valor inicial
        interactive=True
    )
    def generate_pie_chart(*data):
        sections = []
        for value, title, color in data:
            section = ft.PieChartSection(
                value,
                title=title,
                title_style=ft.TextStyle(color="white", size=12),
                color=color,
                radius=60,
            )
            sections.append(section)

        return sections
    def generate_bar_chart(data):
        bar_groups = []
        labels = []
        max_y = 0

        for idx, (month, total_amount) in enumerate(data):
            bar_group = ft.BarChartGroup(
                x=idx,
                bar_rods=[
                    ft.BarChartRod(
                        from_y=0,
                        to_y=total_amount,
                        width=40,
                        #color=ft.Colors.BLUE,
                        tooltip=f"{month}: ${total_amount}",

                        border_radius=0,
                    ),
                ],
            )
            bar_groups.append(bar_group)
            labels.append(ft.ChartAxisLabel(value=idx, label=ft.Container(ft.Text(month), padding=5)))
            if total_amount > max_y:
                max_y = total_amount + 100

        return bar_groups, labels, max_y
    def load_dashboard_data(filter_key='1'):
        nonlocal clients_chart, income_chart
        nonlocal income_data, total_clients, active_clients, pendent_clients
        start_date = None
        match filter_key:
            case '1':
                start_date = datetime.now().date() - timedelta(days=30)
            case '2':
                start_date = datetime.now().date() - timedelta(days=90)
            case '3':
                start_date = datetime.now().date() - timedelta(days=365)
        end_date = datetime.now().date()

        income_data = int(total_income(start_date))
        total_clients = client_count(start_date)
        active_clients = len(search_payment('Pago'))
        pendent_clients = len(search_payment('Vencido'))

        card_income.content.controls[0].controls[2].value = str(income_data)
        card_clients.content.controls[0].controls[2].value = str(total_clients)
        card_active.content.controls[0].controls[2].value = str(active_clients)
        card_pendent.content.controls[0].controls[2].value = str(pendent_clients)

        clients_chart_data = [
        (active_clients, f"{active_clients}%", ft.Colors.BLUE),
        (pendent_clients, f"{pendent_clients}%", ft.Colors.RED)]

        income_chart_data = get_group_of_payments(start_date, end_date)
        bar_groups, labels, max_y = generate_bar_chart(income_chart_data)

        clients_chart.sections = generate_pie_chart(*clients_chart_data)
        income_chart.bar_groups = bar_groups
        income_chart.bottom_axis.labels = labels
        income_chart.max_y = max_y

        page.update()
    load_dashboard_data()

    dashboard_view = ft.Column(scroll=ft.ScrollMode.ALWAYS,height=page.window.height -100,width=page.window.width,expand=True, spacing=10, controls=[
        ft.Row(alignment=ft.MainAxisAlignment.CENTER, controls=[ft.Text("Resumo do mês", size=24, weight="bold", color=ft.Colors.BLUE)]),
        ft.Row(alignment=ft.MainAxisAlignment.START, controls=[ft.Text('Selecionar Período: '), dd]),
        ft.Row(alignment=ft.MainAxisAlignment.SPACE_EVENLY, wrap=True, controls=[card_income, card_clients, card_pendent, card_active]),
        ft.Row(wrap=True, controls=[clients_chart, income_chart])
    ])
    #endregion

    #region CLIENT_VIEW
    actual_page = 0
    showed_clients = 0
    quantity = 20
    total_in_db = client_count()
    async def notification_handler(id):
        if notification_method.value == 'email':  #config['notification_settings']['notification_method'] == 'email':
            email_notification(id)
        elif notification_method.value == 'whatsapp':   #config['notification_settings']['notification_method'] == 'whatsapp':
            await whatsapp_notification(id)
    def check_notifications():
        notifications = []

        # Verificar membresías por vencer
        if config["notification_settings"]["payment_expiration"]:
            expiring = get_expiring_memberships(config["notification_settings"]["notification_days"])
            if expiring:
                for client in expiring:
                    days_left = (parse(client[2]).date() - datetime.now().date()).days
                    notifications.append(
                        ft.Row(
                            controls=[
                                ft.Icon(ft.Icons.WARNING, color=ft.Colors.AMBER, size=24),
                                ft.Text(f"Cliente {client[1]} - Mensalidade vence em {days_left} dias", size=16),
                                ft.Container(expand=True,height=1),
                                ft.IconButton(icon=ft.Icons.SEND, tooltip="Enviar notificação", icon_color=ft.Colors.BLUE, on_click=lambda e, id=client[0]: asyncio.run(notification_handler(id))),
                            ],
                            spacing=10,
                        )
                    )

        # Verificar membresías vencidas
        expired = get_expired_memberships()
        if expired:
            for client in expired:
                days_overdue = (datetime.now().date() - parse(client[2]).date()).days
                notifications.append(
                    ft.Row(
                        controls=[
                            ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED, size=24),
                            ft.Text(f"Cliente {client[1]} - Mensalidade vencida há {days_overdue} dias", size=16),
                            ft.Container(expand=True,height=1),
                            ft.IconButton(icon=ft.Icons.SEND, tooltip="Enviar notificação", icon_color=ft.Colors.BLUE, on_click=lambda e, id=client[0]: asyncio.run(notification_handler(id))),
                        ],
                        spacing=10,
                    )
                )

        # Mostrar las notificaciones en un BottomSheet
        if notifications:
            bottom_sheet = ft.BottomSheet(
                content=ft.Column(
                    controls=[
                        ft.Row(alignment=ft.MainAxisAlignment.CENTER,controls=[ft.Text("Notificações", size=20, weight="bold", color=ft.Colors.BLUE)]),
                        *notifications,
                        ft.Container(expand=True,height=50)
                    ],
                    spacing=10,
                    scroll=ft.ScrollMode.ALWAYS,
                ),
                open=True,
            )
            page.overlay.append(bottom_sheet)
            page.update()
    def load_data():
        nonlocal showed_clients,actual_page,quantity
        showed_clients = actual_page * quantity
        clientes.rows.clear()
        data = show_clients(showed_clients,quantity)
        for client in data:
            clientes.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(str(client[0]))),
                ft.DataCell(ft.Text(str(client[1]))),
                ft.DataCell(ft.Text(str(client[2]))),
                ft.DataCell(ft.Text(str(client[3]))),
                ft.DataCell(ft.Text(str(client[4]))),
                ft.DataCell(ft.Text(str(client[5]))),
                ft.DataCell(ft.Text(str(client[6]))),
                ft.DataCell(ft.Text(str(client[7]))),
                ft.DataCell(ft.Container(border_radius=10,bgcolor='RED' if str(client[8]) == 'Vencido' else 'GREEN',content=ft.Text(' '+str(client[8]+' '),color='WHITE'))),
                ft.DataCell(ft.Row(spacing=0,controls=[
                    ft.IconButton(icon=ft.Icons.EDIT,tooltip='Editar',icon_color=ft.Colors.YELLOW,on_click=lambda e,id=client[0]: edit_client_handler(id)),
                    ft.IconButton(icon=ft.Icons.DELETE,tooltip='Apagar',icon_color=ft.Colors.RED,on_click=lambda e, id=client[0]: del_client_handler(id))])),
                ft.DataCell(ft.Row(alignment=ft.MainAxisAlignment.CENTER,controls=[ft.IconButton(icon=ft.Icons.MONETIZATION_ON,tooltip='Pagamentos',icon_color=ft.Colors.GREEN,on_click=lambda e,id=client[0] :open_payment_modal(id))]))
                ]))
        page.update()
    def regist_payment_handler(e,id ,amount,date):
        date = parse(date,dayfirst=True).date()
        status= 'Pago' if datetime.now().date() <= date + timedelta(days=30) else 'Vencido'
        regist_payment(id,date,amount,status)
        open_payment_modal(id)
        load_data()
    def open_payment_modal(id):
        client = show_client_by_id(id)
        amount_input =ft.TextField(label='Quantidade',width=200, value=config['default_payment'],prefix_text='R$')
        date_input =ft.TextField(label='Data',hint_text='DD/MM/YY',width=200)
        modal = ft.AlertDialog(modal=True, content=ft.Column(spacing=20,controls=[
            ft.Row([ft.Text(f'{client[1]}',size=15,weight='BOLD')],alignment=ft.MainAxisAlignment.CENTER),
            ft.Row([ft.Text(f'Historial de pagamentos'),ft.Container(height=10,expand=True),ft.IconButton(icon=ft.Icons.CLOSE,on_click=lambda e: page.close(modal))]),
            date_input,
            amount_input,
            ft.ElevatedButton(text='Confirmar pagamento',on_click=lambda e: regist_payment_handler(e,client[0],amount_input.value,date_input.value)),
            ft.Column(height=200,scroll=ft.ScrollMode.ALWAYS,controls=[ft.DataTable(
                columns=[
                    ft.DataColumn(ft.Text('Data')),
                    ft.DataColumn(ft.Text('Valor')),
                    ft.DataColumn(ft.Text('Estado'))
                ],
                rows=[
                    ft.DataRow(cells=[
                        ft.DataCell(ft.Text(pago[2])),
                        ft.DataCell(ft.Text(f'${pago[3]}')),
                        ft.DataCell(ft.Text(pago[4]))
                    ]) for pago in show_payments(id)
                ]
            )])

        ]))
        page.open(modal)
    def pagination_handler(e):
        nonlocal actual_page
        if e.control.text == '<' and actual_page == 0:
            back_btn.disabled = True
        if e.control.text == '<' and actual_page > 0:
            if next_btn.disabled:
                next_btn.disabled = False
            actual_page -= 1
        elif e.control.text == '>' and (actual_page +1) * quantity < total_in_db:
            actual_page += 1
            if back_btn.disabled:
                back_btn.disabled = False
        elif e.control.text == '>' and (actual_page +1) * quantity > total_in_db:
            next_btn.disabled = True
        load_data()
    def edit_client_handler(id):
        client = show_client_by_id(id)
        id_input.value = str(client[0])
        nome_input.value = str(client[1])
        number_input.value = str(client[2])
        cpf_input.value = str(client[3])
        email_input.value = str(client[4])
        date_input.value = str(parse(str(client[5]),dayfirst=True).date())
        amount_input.value = str(client[7])
        #logica para editar el usuario
        page.update()
    def del_client_handler(id):
        delete_client(id)
        load_data()
        page.update()
    def search(e):
        clientes.rows.clear()
        filtro = search_input.value
        data = search_db(filtro)
        for client in data:
            clientes.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(str(client[0]))),
                ft.DataCell(ft.Text(str(client[1]))),
                ft.DataCell(ft.Text(str(client[2]))),
                ft.DataCell(ft.Text(str(client[3]))),
                ft.DataCell(ft.Text(str(client[4]))),
                ft.DataCell(ft.Text(str(client[5]))),
                ft.DataCell(ft.Text(str(client[6]))),
                ft.DataCell(ft.Text(str(client[7]))),
                ft.DataCell(ft.Container(border_radius=10,bgcolor='RED' if str(client[8]) == 'Vencido' else 'GREEN',content=ft.Text(' '+str(client[8]+' '),color='WHITE'))),
                ft.DataCell(ft.Row(spacing=0,controls=[
                    ft.IconButton(icon=ft.Icons.EDIT,tooltip='Editar',icon_color=ft.Colors.YELLOW,on_click=lambda e,id=client[0]: edit_client_handler(id)),
                    ft.IconButton(icon=ft.Icons.DELETE,tooltip='Apagar',icon_color=ft.Colors.RED,on_click=lambda e, id=client[0]: del_client_handler(id))
                ])),
                                 ft.DataCell(ft.Row(alignment=ft.MainAxisAlignment.CENTER,controls=[ft.IconButton(icon=ft.Icons.MONETIZATION_ON,tooltip='Pagamentos',icon_color=ft.Colors.GREEN,on_click=lambda e,id=client[0] :open_payment_modal(id))]))
                ]))
        page.update()
    def register_button_clicked(e):
        snack = None
        if any(not input.value.strip() for input in [nome_input,number_input,cpf_input,email_input,date_input,amount_input]):
            snack = ft.SnackBar(content=ft.Text('Todos os campos deben ser preenchidos para poder guardar o cliente.',color=ft.Colors.WHITE),bgcolor=ft.Colors.RED)
        else:
            if client_exist(id_input.value):
                edit_client(id_input.value,nome_input.value,number_input.value,cpf_input.value,email_input.value,date_input.value,amount_input.value)
                snack = ft.SnackBar(content=ft.Text('Se editou o cliente corretamente.',color=ft.Colors.BLACK),bgcolor=ft.Colors.YELLOW)
            else:
                regist_client(nome_input.value,number_input.value,cpf_input.value,email_input.value,date_input.value,amount_input.value)
                snack = ft.SnackBar(content=ft.Text('Se cadastrou o cliente corretamente.',color=ft.Colors.WHITE),bgcolor=ft.Colors.GREEN)

            id_input.value =''
            nome_input.value=''
            number_input.value=''
            cpf_input.value=''
            email_input.value =''
            date_input.value =''
            amount_input.value = ''

        page.overlay.append(snack)
        snack.open = True
        load_data()
        page.update()
    clientes = ft.DataTable(scale=1,column_spacing=10,border=ft.border.all(2,'BLACK'),border_radius=10,
        columns=[
                ft.DataColumn(ft.Text('ID'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Nome'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Numero'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('CPF'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Email'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Data de inicio'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Data de vencimento'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Valor da mensalidade'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Estado'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Ações'),heading_row_alignment=ft.MainAxisAlignment.CENTER),
                ft.DataColumn(ft.Text('Pagamentos'),heading_row_alignment=ft.MainAxisAlignment.CENTER)
            ])
    id_input = ft.TextField(visible=False)
    nome_input= ft.TextField(label='Nome',width=150,height=40,text_style=ft.TextStyle(size=14))
    number_input= ft.TextField(label='Numero',width=150,height=40,text_style=ft.TextStyle(size=14))
    cpf_input= ft.TextField(label='CPF',width=150,height=40,text_style=ft.TextStyle(size=14))
    email_input= ft.TextField(label='Email',width=150,height=40,text_style=ft.TextStyle(size=14))
    date_input= ft.TextField(label='Data de inicio',width=150,height=40,hint_text='DD/MM/YY',text_style=ft.TextStyle(size=14))
    amount_input = ft.TextField(label='Valor da mensalidade',width=150,height=40,text_style=ft.TextStyle(size=14))
    search_input =ft.TextField(height=40,text_style=ft.TextStyle(size=14),label='Buscar',on_change=search)

    responsive_column =ft.Column(scroll=ft.ScrollMode.ALWAYS,controls=[clientes])
    responsive_row =ft.Row(spacing=0,scroll=ft.ScrollMode.ALWAYS,controls=[responsive_column])
    
    back_btn =ft.ElevatedButton(text='<',on_click=pagination_handler,disabled=True)
    next_btn =ft.ElevatedButton(text='>',on_click=pagination_handler)
    notifications_btn = ft.IconButton(icon_size=24,icon=ft.Icons.NOTIFICATIONS,icon_color=ft.Colors.BLUE,on_click=lambda e: check_notifications())
    clients_view = ft.Column(spacing=10, controls=[
        ft.Row(alignment=ft.MainAxisAlignment.CENTER, controls=[ft.Text("Clientes", size=24, weight="bold", color=ft.Colors.BLUE)]),
        ft.Row(wrap=True, controls=[search_input]),
        ft.Text("Cadastrar clientes",size=14,weight="bold",color=ft.Colors.BLUE),
        ft.Row(wrap=True, controls=[id_input, nome_input, number_input,cpf_input, email_input, date_input, amount_input, ft.ElevatedButton(text='Salvar', color=ft.Colors.WHITE, bgcolor=ft.Colors.GREEN, on_click=register_button_clicked)]),
        ft.Row(controls=[back_btn, next_btn,ft.Container(expand=True,height=1),notifications_btn,ft.Container(width=55,height=1)]),
        responsive_row
    ])
    #endregion

    #region SETTINGS VIEW

    def import_excel(e, file_picker: ft.FilePicker):
        try:
            file_path = file_picker.result.files[0].path
            ws = load_workbook(file_path).active
        
            for row in ws.iter_rows(min_row=2):
                if len(row) >= 5:
                    regist_client(
                        name=row[0],
                        number=row[1],
                        cpf=row[2],
                        email=row[3],
                        start_date=row[4],
                        amount=row[5]
                    )
                else:
                    snack = ft.SnackBar(content=ft.Text(f"Alerta: O Arquivo deve ter pelo menos 6 colunas (ler a ajuda para mais informação"),bgcolor=ft.Colors.YELLOW)
                    page.overlay.append(snack)
                    snack.open = True
 

            #mensaje de confirmacion
            snack =ft.SnackBar(content=ft.Text("Dados importados com sucesso!"),bgcolor=ft.Colors.GREEN)
            page.overlay.append(snack)
            snack.open = True

            #mensaje de error
        except Exception:
            snack = ft.SnackBar(content=ft.Text(f"Erro: Não se importaram os dados"),bgcolor=ft.Colors.RED)
            page.overlay.append(snack)
            snack.open = True
        page.update()

    def save_settings(e):
        config = {
            "theme": theme_dd.value,
            "default_payment": float(default_payment.value),
            "language": language_dd.value,
            "notification_settings": {
                "payment_expiration": notifications_expiration.value,
                "notification_days": int(notification_days.value),
                "notification_method": notification_method.value,
                "due_msg":str(due_msg),
                "overdue_msg":str(overdue_msg)
            },
            "bussines_name":str(bussines_name.value),
            "pix_key":str(pix_key.value)
        }
        save_config(config)
        snack = ft.SnackBar(content=ft.Text("Configurações salvas com sucesso!"),bgcolor=ft.Colors.GREEN)
        page.overlay.append(snack)
        snack.open = True
        page.update()

    file_picker = ft.FilePicker(on_result=lambda e: import_excel(e, file_picker))
    page.overlay.append(file_picker)
    theme_dd = ft.Dropdown(value=config["theme"],label="Tema do Sistema",width=200,
        options=[
            ft.dropdown.Option("light", "Claro"),
            ft.dropdown.Option("dark", "Escuro")
        ],
        on_change=lambda e: theme_changed(e)
    )
    default_payment = ft.TextField(label="Valor padrão da mensalidade",value=str(config["default_payment"]),width=200,prefix_text="R$",on_change=lambda e: save_settings(e),)
    language_dd = ft.Dropdown(value=config["language"],label="Idioma",width=200,
        options=[
            ft.dropdown.Option("pt", "Português"),
            ft.dropdown.Option("en", "English"),
            ft.dropdown.Option("es", "Español")
        ],
        disabled=True,
        on_change=lambda e: save_settings(e)
    )
    notifications_expiration = ft.Switch(label="Notificar vencimento",value=config["notification_settings"]["payment_expiration"],on_change=lambda e: save_settings(e))
    notification_days = ft.TextField(label="Dias de antecedência a notificar",value=str(config["notification_settings"]["notification_days"]),width=200,on_change=lambda e: save_settings(e))
    notification_method = ft.Dropdown(value=config['notification_settings']['notification_method'],label='Metodo de notificação',width=200,
        options=[
            ft.dropdown.Option('email','email'),
            ft.dropdown.Option('whatsapp','whatsapp(Beta)')
        ],
        on_change=lambda e: save_settings(e))
    due_msg = config['notification_settings']['due_msg']
    overdue_msg = config['notification_settings']['overdue_msg']
    def edit_text_modal(e):
        nonlocal due_msg,overdue_msg
        def switch_value(e):
            nonlocal due_msg,overdue_msg
            dialog_edit.content.controls[1].value = str(overdue_msg) if e.control.value == False else str(due_msg)
            page.update()
        def save(e):
            nonlocal due_msg,overdue_msg
            if switch.value:
                due_msg = e.control.value   
            else:
                overdue_msg = e.control.value

        switch = ft.Switch(on_change=switch_value)
        dialog_edit = ft.AlertDialog(modal=True,title=ft.Text('Modificar texto de notificação',size=24,weight='bold',color=ft.Colors.BLUE),content=ft.Column(width=400,height=500,controls=[
            ft.Row(controls=[ft.Text('Pre-Vencimento'),switch,ft.Text('Vencimento')]),
            ft.TextField(multiline=True,expand=True,value=overdue_msg,on_change=lambda e: save(e))
        ]),
        actions=[ft.TextButton(text="Cerrar",on_click=lambda e: page.close(dialog_edit)),
                 ft.FilledButton(text='Salvar',bgcolor=ft.Colors.GREEN,on_click=lambda e: save_settings(e))]
        )
        page.open(dialog_edit)

    due_btn = ft.FilledButton(text='Customizar',on_click=edit_text_modal)
    bussines_name = ft.TextField(label="Nome do Negocio",value=config["bussines_name"],width=200,on_change=lambda e: save_settings(e))
    pix_key = ft.TextField(label="Chave Pix",value=config["pix_key"],width=200,on_change=lambda e: save_settings(e))
   
    def show_help(e):
        dialog_help = ft.AlertDialog(modal=True,title=ft.Text("Formatação do arquivo Excel", size=24, weight="bold", color=ft.Colors.BLUE),content=ft.Column(width=400,height=500,controls=[ft.Text("Para que a importação aconteça com normalidade o arquivo Excel deve cumprir os sgtes requisitos",size=20),
            ft.Text("* Pelo menos 6 colunas.",size=20),
            ft.Text("* Uma coluna para cada um dos seguintes encabezados (Nome,Numero,CPF,Email,Data_Inicio,Quantidade) referente a os dados dessa coluna.",size=20),
            ft.Text("Esta funcionalidade ainda esta em desenvolvimento , erros de importação podem acontecer (Salve a gym.db antes de fazer esta operação)",color=ft.Colors.RED,weight="bold",size=15)
            ]),
            actions=[ft.TextButton(text="Entendido",on_click=lambda e: page.close(dialog_help))])      
        page.open(dialog_help)
        
    settings_view = ft.Column(
        spacing=20,
        height=page.window.height - 100,
        scroll=ft.ScrollMode.ALWAYS,
        controls=[
            ft.Row(alignment=ft.MainAxisAlignment.CENTER, controls=[ft.Text("Configurações do sistema", size=24, weight="bold", color=ft.Colors.BLUE)]),
            ft.Row(wrap=True,controls=[
            # Importação de dados
            ft.Container(
                content=ft.Column([
                    ft.Text("Importar Dados", size=18, weight="bold"),
                        ft.Row([
                        ft.ElevatedButton(
                            "Importar do Excel",
                            icon=ft.Icons.FILE_UPLOAD,
                                on_click=lambda _:file_picker.pick_files(allowed_extensions=['xlsx','xls'])
                            ),
                            ft.IconButton(icon=ft.Icons.HELP,tooltip="Ajuda",icon_color = ft.Colors.CYAN,icon_size = 20, on_click=show_help)
                    ])
                ]),
                padding=20,
                border=ft.border.all(1, ft.Colors.GREY_400),
                border_radius=10,
                height=300,
                width=300
            ),
            # Configurações do sistema
            ft.Container(
                content=ft.Column([
                    ft.Text("Preferências do Sistema", size=18, weight="bold"),
                    theme_dd,
                    language_dd,
                ]),
                padding=20,
                border=ft.border.all(1, ft.Colors.GREY_400),
                border_radius=10,
                height=300,
                width=300
            ),
            # Configurações de pagamento
            ft.Container(
                content=ft.Column([
                    ft.Text("Configurações de Pagamento", size=18, weight="bold"),
                    default_payment,
                ]),
                padding=20,
                border=ft.border.all(1, ft.Colors.GREY_400),
                border_radius=10,
                height=300,
                width=300
            ),
            # Configurações de notificação
            ft.Container(
                content=ft.Column([
                    ft.Text("Notificações", size=18, weight="bold"),
                    notifications_expiration,
                    notification_days,
                    notification_method,
                    due_btn,
                ]),
                padding=20,
                border=ft.border.all(1, ft.Colors.GREY_400),
                border_radius=10,
                height=300,
                width=300
            ),
            # Configurações do negocio
            ft.Container(
                content=ft.Column([
                    ft.Text("Dados do Negocio", size=18, weight="bold"),
                    bussines_name,
                    pix_key
                ]),
                padding=20,
                border=ft.border.all(1, ft.Colors.GREY_400),
                border_radius=10,
                height=300,
                width=300
            )
            ]),
        ]
    )
    #endregion

    #region ABOUT VIEW
    about_view = ft.Column(
        expand=True,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        controls=[
            ft.Text("SimpleGym", size=32, weight="bold", color=ft.Colors.BLUE),
            ft.Text(
                "A solução completa para a gestão de academias.",
                size=18,
                italic=True,
                color=ft.Colors.GREY_600,
            ),
            ft.Divider(height=20, thickness=1, color=ft.Colors.GREY_400),
            ft.Text(
                "Principais características:",
                size=20,
                weight="bold",
                color=ft.Colors.BLUE,
            ),
            ft.ListView(
                expand=False,
                spacing=10,
                controls=[
                    ft.Row(
                        controls=[
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN),
                            ft.Text("Gestão de clientes e pagamentos", size=16),
                        ]
                    ),
                    ft.Row(
                        controls=[
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN),
                            ft.Text("Relatórios financeiros", size=16),
                        ]
                    ),
                    ft.Row(
                        controls=[
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN),
                            ft.Text("Interface intuitiva e fácil de usar", size=16),
                        ]
                    ),
                    ft.Row(
                        controls=[
                            ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN),
                            ft.Text("Notificações via E-mail e Whatsapp(beta)", size=16),
                        ]
                    ),
                ],
            ),
            ft.Text(
                "Atualizações recentes:",
                size=20,
                weight="bold",
                color=ft.Colors.BLUE,
            ),
            ft.Row(wrap=True,controls=[
                ft.Column(controls=[
                    ft.Text("Adição da notificação via whatsapp."),
                    ft.Text("Posibilidade de customizar o message de notificação.")
                ]),
                ft.Column(controls=[
                    ft.Text("Agora ao cadastar um cliente , tambem cadastra um pagamento com a mesma data e valor"),
                    ft.Text("Correção de falhas como scrolls, tamanhos de letras e interação con a base de dados")
                ])
            ]),
            ft.Divider(height=20, thickness=1, color=ft.Colors.GREY_400),
            ft.Text("Desenvolvido por: Freyre Corona", size=16, color=ft.Colors.GREY_700),
            ft.Text("Versão: 1.2.0", size=16, color=ft.Colors.GREY_700),
        ],
    )
    #endregion

    # Navigation Rail e manipulador de eventos
    def change_destination(e):
        index = e.control.selected_index
        content.controls.clear()
        match index:
            case 0:
                content.controls.append(dashboard_view)
                load_dashboard_data()
            case 1:
                content.controls.append(clients_view)
                load_data()
            case 2:
                content.controls.append(settings_view)
            case 3:
                content.controls.append(about_view)
        page.update()

    rail = ft.NavigationRail(
        selected_index=0,
        label_type=ft.NavigationRailLabelType.ALL,
        min_width=50,
        min_extended_width=100,
        destinations=[
            ft.NavigationRailDestination(label='Resumo', icon=ft.Icons.DASHBOARD),
            ft.NavigationRailDestination(label='Clientes', icon=ft.Icons.GROUPS),
            ft.NavigationRailDestination(label='Configurações', icon=ft.Icons.SETTINGS),
            ft.NavigationRailDestination(label='Sobre nós...', icon=ft.Icons.INFO)
        ],
        on_change=change_destination
    )

    content = ft.Column([dashboard_view], expand=True)
    resized_window(None)
    page.add(ft.Row(expand=True, controls=[rail, ft.VerticalDivider(width=1), content]))

# Iniciar la app
ft.app(target=main)
